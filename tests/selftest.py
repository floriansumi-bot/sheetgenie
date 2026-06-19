"""Offline self-tests for SheetGenie's serverless functions.

Runs without any network or API key:
  * generate.py  — real openpyxl round-trip (formulas, charts, injection defuse,
                   validation limits, edge cases).
  * improve.py   — the do_POST flow with a mocked Anthropic client (success,
                   model fallback chain, max_tokens, missing key, bad input).

Run:  .venv\\Scripts\\python.exe tests\\selftest.py
Exit code 0 = all passed.
"""

import importlib.util
import io
import json
import os
import sys
import types

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _load(name, relpath):
    path = os.path.join(ROOT, relpath)
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


PASSED = 0
FAILED = 0


def check(label, cond):
    global PASSED, FAILED
    if cond:
        PASSED += 1
        print("  PASS  " + label)
    else:
        FAILED += 1
        print("  FAIL  " + label)


# ===========================================================================
# generate.py
# ===========================================================================
def test_generate():
    print("\n[generate.py]")
    from openpyxl import load_workbook

    gen = _load("generate_mod", os.path.join("api", "generate.py"))

    # --- SPEC.md worked example: budget tracker -----------------------------
    budget = {
        "title": "Monthly Budget Tracker",
        "sheets": [{
            "name": "Budget",
            "columns": [
                {"header": "Category", "type": "text"},
                {"header": "Budgeted", "type": "currency"},
                {"header": "Actual", "type": "currency"},
                {"header": "Variance", "type": "formula", "formula": "=B{row}-C{row}"},
            ],
            "rows": [
                ["Rent", 1500, 1500, None],
                ["Groceries", 400, 462.30, None],
                ["Transport", 120, 98.5, None],
            ],
            "charts": [{
                "type": "bar", "title": "Budgeted vs Actual",
                "categoriesColumn": 1, "valueColumns": [2, 3],
                "dataEndRow": 99,  # deliberately over-specified -> must clamp
            }],
        }],
    }
    gen._validate_spec(budget)
    data = gen._build_workbook(budget)
    check("budget: produces non-empty xlsx bytes", isinstance(data, bytes) and len(data) > 0)

    wb = load_workbook(io.BytesIO(data))
    ws = wb["Budget"]
    check("budget: sheet named 'Budget'", ws.title == "Budget")
    check("budget: header A1 == 'Category'", ws["A1"].value == "Category")
    check("budget: header is bold", bool(ws["A1"].font and ws["A1"].font.bold))
    check("budget: D2 is a real formula =B2-C2", ws["D2"].data_type == "f" and ws["D2"].value == "=B2-C2")
    check("budget: D4 formula row-substituted =B4-C4", ws["D4"].value == "=B4-C4")
    check("budget: currency B2 number_format applied", ws["B2"].number_format == "#,##0.00")
    check("budget: freeze panes at A2", ws.freeze_panes == "A2")
    check("budget: one chart present", len(ws._charts) == 1)
    # Chart series must be clamped to the 3 populated rows, not row 99.
    ch = ws._charts[0]
    refs = []
    for s in ch.series:
        try:
            refs.append(s.val.numRef.f)
        except Exception:
            pass
    check("budget: chart series clamped to row 4 (not 99)", all("99" not in r for r in refs) and any("$4" in r or "4" in r for r in refs))

    # --- Column auto-sizing: content-aware widths + long-text wrap ----------
    sized = {
        "title": "sz",
        "sheets": [{
            "name": "S",
            "columns": [
                {"header": "ID", "type": "number"},          # short header + short data
                {"header": "Name", "type": "text"},          # data longer than the 4-char header
                {"header": "Description", "type": "text"},    # very long data -> cap + wrap
                {"header": "Price", "type": "currency"},
            ],
            "rows": [
                [1, "Alexander Hamilton", "x" * 120, 1234.5],
                [2, "Bob", "short", 9.99],
            ],
            "autoFilter": True,
        }],
    }
    gen._validate_spec(sized)
    wbS = load_workbook(io.BytesIO(gen._build_workbook(sized))).active
    wID = wbS.column_dimensions["A"].width
    wName = wbS.column_dimensions["B"].width
    wDesc = wbS.column_dimensions["C"].width
    check("width: short numeric column stays narrow (< 14)", wID is not None and wID < 14)
    check("width: name column sized to its data, not the short header", wName >= len("Alexander Hamilton"))
    check("width: long-text column capped at MAX", wDesc <= gen.MAX_COL_WIDTH)
    check("width: long-text data cell wraps", bool(wbS["C2"].alignment and wbS["C2"].alignment.wrap_text))
    check("width: normal-text column does not wrap", not (wbS["B2"].alignment and wbS["B2"].alignment.wrap_text))

    # a SANE explicit spec width is honoured over the auto measurement
    expl = {"title": "e", "sheets": [{"name": "S", "columns": [{"header": "A", "type": "text", "width": 33}], "rows": [["hi"]]}]}
    gen._validate_spec(expl)
    check("width: sane explicit spec width honoured",
          load_workbook(io.BytesIO(gen._build_workbook(expl))).active.column_dimensions["A"].width == 33.0)

    # an ABSURD explicit width (model sometimes emits 150-250) is rejected -> sane auto width
    huge = {"title": "h", "sheets": [{"name": "S", "columns": [{"header": "Notes", "type": "text", "width": 250}], "rows": [["short note"]]}]}
    gen._validate_spec(huge)
    wHuge = load_workbook(io.BytesIO(gen._build_workbook(huge))).active.column_dimensions["A"].width
    check("width: absurd explicit width (250) ignored -> <= MAX", wHuge <= gen.MAX_COL_WIDTH)

    # _auto_widths unit behaviour
    aw, wrap = gen._auto_widths(sized["sheets"][0]["columns"], sized["sheets"][0]["rows"], True)
    check("auto_widths: one clamped width per column",
          len(aw) == 4 and all(gen.MIN_COL_WIDTH <= w <= gen.MAX_COL_WIDTH for w in aw))
    check("auto_widths: flags the long-text column for wrap", wrap == {2})

    # --- Formula injection defuse (non-formula text column) -----------------
    inj = {
        "title": "inj",
        "sheets": [{
            "name": "S",
            "columns": [{"header": "Note", "type": "text"}],
            "rows": [["=HYPERLINK(\"http://evil\",\"x\")"], ["+danger"], ["normal"]],
        }],
    }
    gen._validate_spec(inj)
    wb2 = load_workbook(io.BytesIO(gen._build_workbook(inj)))
    s2 = wb2.active
    check("injection: leading-= cell stored as string, not formula", s2["A2"].data_type == "s")
    check("injection: value preserved literally (no apostrophe)", s2["A2"].value == "=HYPERLINK(\"http://evil\",\"x\")")
    check("injection: '+danger' is plain text", s2["A3"].data_type == "s" and s2["A3"].value == "+danger")

    # --- Empty template (no rows -> no chart, no crash) ---------------------
    empty = {"title": "t", "sheets": [{"name": "S", "columns": [{"header": "A", "type": "text"}], "rows": [],
                                       "charts": [{"type": "bar", "title": "c", "categoriesColumn": 1, "valueColumns": [1]}]}]}
    gen._validate_spec(empty)
    wb3 = load_workbook(io.BytesIO(gen._build_workbook(empty)))
    check("empty: builds with zero data rows", wb3.active["A1"].value == "A")
    check("empty: no chart when no data rows", len(wb3.active._charts) == 0)

    # --- date coercion + pie chart ------------------------------------------
    dated = {"title": "d", "sheets": [{"name": "S",
             "columns": [{"header": "Day", "type": "date"}, {"header": "N", "type": "number"}],
             "rows": [["2026-01-15", 5], ["2026-02-20", 9]],
             "charts": [{"type": "pie", "title": "p", "categoriesColumn": 1, "valueColumns": [2]}]}]}
    gen._validate_spec(dated)
    wb4 = load_workbook(io.BytesIO(gen._build_workbook(dated)))
    import datetime as _dt
    check("date: ISO string coerced to a date", isinstance(wb4.active["A2"].value, (_dt.date, _dt.datetime)))
    check("date: pie chart present", len(wb4.active._charts) == 1)

    # --- Chart series/category alignment with dataStartRow > 2 (regression) ---
    import re as _re
    aligned = {"title": "a", "sheets": [{"name": "S",
               "columns": [{"header": "Label", "type": "text"}, {"header": "Val", "type": "number"}],
               "rows": [["r2", 10], ["r3", 20], ["r4", 30], ["r5", 40], ["r6", 50]],
               "charts": [{"type": "bar", "title": "c", "categoriesColumn": 1,
                           "valueColumns": [2], "dataStartRow": 3}]}]}
    gen._validate_spec(aligned)
    chA = load_workbook(io.BytesIO(gen._build_workbook(aligned))).active._charts[0]

    def _rows(f):
        m = _re.findall(r"\$[A-Z]+\$(\d+)", f or "")
        return (int(m[0]), int(m[-1])) if m else None

    ser = chA.series[0]
    val_rows = _rows(ser.val.numRef.f)
    cat_ref = (ser.cat.strRef.f if ser.cat and ser.cat.strRef
               else (ser.cat.numRef.f if ser.cat and ser.cat.numRef else None))
    cat_rows = _rows(cat_ref)
    check("chart align: series span == category span (dataStartRow=3)", val_rows == cat_rows)
    check("chart align: both span rows 3..6", val_rows == (3, 6))

    # --- Advanced features: totals row, conditional format, validation, named range ---
    adv = {
        "title": "adv",
        "namedRanges": [{"name": "TaxRate", "ref": "'S'!$B$1"}],
        "sheets": [{
            "name": "S",
            "columns": [
                {"header": "Item", "type": "text"},
                {"header": "Amount", "type": "currency"},
                {"header": "Status", "type": "text"},
            ],
            "rows": [["A", 600, "To Do"], ["B", 200, "Done"], ["C", 900, "In Progress"]],
            "totalsRow": True,
            "conditionalFormats": [{"column": 2, "rule": "greaterThan", "value": 500, "color": "red"}],
            "dataValidations": [{"column": 3, "values": ["To Do", "In Progress", "Done"]}],
        }],
    }
    gen._validate_spec(adv)
    wbA = load_workbook(io.BytesIO(gen._build_workbook(adv)))
    wsA = wbA["S"]
    cellsA = [c for row in wsA.iter_rows() for c in row]
    has_sum = any(c.data_type == "f" and "SUM" in str(c.value).upper() for c in cellsA)
    has_total_label = any(str(c.value or "").strip().lower() == "total" for c in cellsA)
    check("totals: a live =SUM formula is present", has_sum)
    check("totals: a 'Total' label is present", has_total_label)
    check("condfmt: a rule exists on the sheet", len(list(wsA.conditional_formatting)) > 0)
    check("validation: a list validation exists", len(wsA.data_validations.dataValidation) > 0)
    check("namedRange: a defined name is present", len(list(wbA.defined_names)) > 0)

    def rejects_adv(label, spec):
        try:
            gen._validate_spec(spec)
            check(label + " (should reject)", False)
        except gen.SpecError:
            check(label, True)

    rejects_adv("reject: invalid condfmt rule", {"title": "t", "sheets": [{"name": "S",
                "columns": [{"header": "a", "type": "number"}], "rows": [[1]],
                "conditionalFormats": [{"column": 1, "rule": "wat", "value": 1}]}]})
    rejects_adv("reject: condfmt column out of range", {"title": "t", "sheets": [{"name": "S",
                "columns": [{"header": "a", "type": "number"}], "rows": [[1]],
                "conditionalFormats": [{"column": 5, "rule": "greaterThan", "value": 1}]}]})
    rejects_adv("reject: validation column out of range", {"title": "t", "sheets": [{"name": "S",
                "columns": [{"header": "a", "type": "text"}], "rows": [["x"]],
                "dataValidations": [{"column": 9, "values": ["a"]}]}]})
    rejects_adv("reject: dropdown list over 255 chars", {"title": "t", "sheets": [{"name": "S",
                "columns": [{"header": "a", "type": "text"}], "rows": [["x"]],
                "dataValidations": [{"column": 1, "values": ["opt" + str(i) + "x" * 20 for i in range(15)]}]}]})
    rejects_adv("reject: illegal named-range name (spaces)", {"title": "t",
                "namedRanges": [{"name": "Bad Name", "ref": "'S'!$A$1"}],
                "sheets": [{"name": "S", "columns": [{"header": "a", "type": "text"}], "rows": [["x"]]}]})

    # dropdown list escapes embedded double-quotes (no formula break-out)
    quoted = {"title": "q", "sheets": [{"name": "S",
              "columns": [{"header": "Pick", "type": "text"}], "rows": [["x"], ["y"]],
              "dataValidations": [{"column": 1, "values": ['Hi "Bob"', "Plain"]}]}]}
    gen._validate_spec(quoted)
    wbQ = load_workbook(io.BytesIO(gen._build_workbook(quoted)))
    dvf = wbQ["S"].data_validations.dataValidation[0].formula1
    check("validation: embedded quotes doubled (even parity, no breakout)",
          '""' in dvf and dvf.count('"') % 2 == 0)

    # --- Validation rejections (expect SpecError -> 400) --------------------
    def rejects(label, spec):
        try:
            gen._validate_spec(spec)
            check(label + " (should reject)", False)
        except gen.SpecError:
            check(label, True)

    rejects("reject: 0 sheets", {"title": "t", "sheets": []})
    rejects("reject: 9 sheets", {"title": "t", "sheets": [{"name": "S", "columns": [{"header": "a", "type": "text"}]}] * 9})
    rejects("reject: bad column type", {"title": "t", "sheets": [{"name": "S", "columns": [{"header": "a", "type": "wat"}]}]})
    rejects("reject: formula column without formula", {"title": "t", "sheets": [{"name": "S", "columns": [{"header": "a", "type": "formula"}]}]})
    rejects("reject: chart col index out of range", {"title": "t", "sheets": [{"name": "S", "columns": [{"header": "a", "type": "text"}],
            "charts": [{"type": "bar", "title": "c", "categoriesColumn": 5, "valueColumns": [1]}]}]})
    rejects("reject: oversized cell (>32767 chars)", {"title": "t", "sheets": [{"name": "S", "columns": [{"header": "a", "type": "text"}],
            "rows": [["x" * 40000]]}]})
    rejects("reject: too many total cells", {"title": "t", "sheets": [{"name": "S",
            "columns": [{"header": "c%d" % i, "type": "number"} for i in range(50)],
            "rows": [[1] * 50 for _ in range(5000)]}]})

    # --- filename slugify ----------------------------------------------------
    check("slug: from title", gen._slugify_filename(None, {"title": "My Q4 Report!!"}) == "my-q4-report.xlsx")
    check("slug: fallback", gen._slugify_filename("", {}) == "spreadsheet.xlsx")
    check("slug: no path traversal", "/" not in gen._slugify_filename("../../etc/passwd", {}) and "\\" not in gen._slugify_filename("..\\x", {}))


# ===========================================================================
# improve.py  (mocked Anthropic client)
# ===========================================================================
CANNED = {
    "status": "ready",
    "improvedPrompt": "A monthly budget tracker with category, budgeted, actual, variance.",
    "notes": "Made you a budget tracker.",
    "spec": {"title": "Budget", "sheets": [{"name": "Budget",
             "columns": [{"header": "Category", "type": "text"}], "rows": [["Rent"]]}]},
}


def _drive(imp, body_bytes, content_length=None, env=None):
    """Invoke imp.handler.do_POST with a fake request; capture (status, json)."""
    captured = {"status": None, "headers": {}, "body": b""}

    class TestHandler(imp.handler):
        def __init__(self):  # bypass BaseHTTPRequestHandler socket setup
            self.rfile = io.BytesIO(body_bytes)
            self.wfile = io.BytesIO()
            cl = content_length if content_length is not None else len(body_bytes)
            self.headers = {"Content-Length": str(cl)}

        def send_response(self, code):
            captured["status"] = code

        def send_header(self, k, v):
            captured["headers"][k] = v

        def end_headers(self):
            pass

    h = TestHandler()
    h.do_POST()

    raw = h.wfile.getvalue()
    try:
        parsed = json.loads(raw.decode("utf-8")) if raw else None
    except Exception:
        parsed = None
    return captured["status"], parsed, h.messages_ref if hasattr(h, "messages_ref") else None


def _set_keys(imp, gemini="g-key", groq=None, xai=None):
    imp.GEMINI_API_KEY = gemini
    imp.GROQ_API_KEY = groq
    imp.XAI_API_KEY = xai


def test_improve():
    print("\n[improve.py]")
    imp = _load("improve_mod", os.path.join("api", "improve.py"))
    _set_keys(imp)  # configured (Gemini key present) by default
    real_generate = imp._generate  # keep a handle to the real chain runner

    def gen_returns(text, store=None):
        def _g(system, user_text, files):
            if store is not None:
                store.update(system=system, user_text=user_text, files=files)
            return text
        return _g

    def gen_raises(reason):
        def _g(system, user_text, files):
            raise imp._ProviderError(reason)
        return _g

    img_b64 = "aGVsbG8="  # dummy base64

    # success
    imp._generate = gen_returns(json.dumps(CANNED))
    status, parsed, _ = _drive(imp, json.dumps({"prompt": "budget tracker"}).encode())
    check("success: 200", status == 200)
    check("success: returns improvedPrompt+notes+spec",
          isinstance(parsed, dict) and {"improvedPrompt", "notes", "spec"} <= set(parsed))

    # needs_input passthrough
    NEEDS = {"status": "needs_input", "notes": "need a couple details",
             "questions": [{"question": "What period?", "hint": "e.g. 2026"}]}
    imp._generate = gen_returns(json.dumps(NEEDS))
    status, parsed, _ = _drive(imp, json.dumps({"prompt": "a report"}).encode())
    check("needs_input: 200 passthrough with status", status == 200 and parsed.get("status") == "needs_input")
    check("needs_input: questions present", isinstance(parsed.get("questions"), list) and len(parsed["questions"]) == 1)

    # config: provider chain default is gemini -> groq (both free)
    check("config: provider chain is gemini then groq", imp.PROVIDER_CHAIN == ["gemini", "groq"])

    # provider fallthrough: when gemini fails, the free Groq fallback answers
    _sc, _sg, _sq = imp.PROVIDER_CHAIN, imp._call_gemini, imp._call_groq
    def _raise_rl(s, u, f): raise imp._ProviderError("rate_limit")
    imp.PROVIDER_CHAIN = ["gemini", "groq"]
    imp._call_gemini = _raise_rl
    imp._call_groq = lambda s, u, f: "GROQ-OUT"
    check("fallthrough: groq answers when gemini is rate-limited", real_generate("s", "u", []) == "GROQ-OUT")
    imp.PROVIDER_CHAIN, imp._call_gemini, imp._call_groq = _sc, _sg, _sq

    # _call_groq with no key is skipped (no_key) so the chain falls through cleanly
    _gk = imp.GROQ_API_KEY; imp.GROQ_API_KEY = None
    try:
        imp._call_groq("s", "u", []); check("groq: missing key raises", False)
    except imp._ProviderError as e:
        check("groq: missing key -> no_key", e.reason == "no_key")
    imp.GROQ_API_KEY = _gk

    # only a Groq key configured (no Gemini) is still a valid configuration
    _set_keys(imp, gemini=None, groq="q-key", xai=None)
    status, parsed, _ = _drive(imp, json.dumps({"prompt": "x"}).encode())
    check("config: groq-only key is configured (not 500-not-configured)",
          not (status == 500 and "not configured" in (parsed or {}).get("error", "")))

    # no provider key at all -> 500
    _set_keys(imp, gemini=None, groq=None, xai=None)
    status, parsed, _ = _drive(imp, json.dumps({"prompt": "x"}).encode())
    check("no key: 500", status == 500)
    check("no key: 'not configured' message", "not configured" in (parsed or {}).get("error", ""))
    _set_keys(imp)  # restore

    # empty prompt -> 400
    status, parsed, _ = _drive(imp, json.dumps({"prompt": "   "}).encode())
    check("empty prompt: 400", status == 400)

    # malformed request JSON -> 400
    status, parsed, _ = _drive(imp, b"{not json")
    check("bad json: 400", status == 400)

    # oversized body -> 413 (Content-Length over the cap)
    status, parsed, _ = _drive(imp, b"{}", content_length=imp.MAX_BODY_BYTES + 1)
    check("oversized: 413", status == 413)

    # clarifications are folded into the prompt sent to the provider
    store = {}
    imp._generate = gen_returns(json.dumps(CANNED), store)
    _drive(imp, json.dumps({"prompt": "sales report",
                            "clarifications": [{"question": "Period?", "answer": "Q1 2026"}]}).encode())
    check("clarifications: folded into the user message",
          "Answers to your questions" in store.get("user_text", "") and "Q1 2026" in store.get("user_text", ""))

    # baseSpec (edit mode) folded in
    store = {}
    imp._generate = gen_returns(json.dumps(CANNED), store)
    _drive(imp, json.dumps({"prompt": "add a Tax column",
                            "baseSpec": {"title": "T", "sheets": [{"name": "S",
                                         "columns": [{"header": "A", "type": "text"}]}]}}).encode())
    check("baseSpec: folded into the user message as an edit instruction",
          "CURRENT SPREADSHEET to edit" in store.get("user_text", "") and "add a Tax column" in store.get("user_text", ""))

    # baseSpec with empty sheets still treated as an edit
    store = {}
    imp._generate = gen_returns(json.dumps(CANNED), store)
    _drive(imp, json.dumps({"prompt": "add a column", "baseSpec": {"title": "T", "sheets": []}}).encode())
    check("baseSpec: empty-sheets spec still framed as an edit",
          "CURRENT SPREADSHEET to edit" in store.get("user_text", ""))

    # chosenLayout (the user picked a proposed layout) folded in as a build instruction
    store = {}
    imp._generate = gen_returns(json.dumps(CANNED), store)
    _drive(imp, json.dumps({"prompt": "a budget",
                            "chosenLayout": {"title": "Monthly Tabs",
                                             "sheets": [{"name": "January", "columns": ["Item", "Amount"]}]}}).encode())
    check("chosenLayout: folded into the user message as a build instruction",
          "CHOSEN LAYOUT" in store.get("user_text", "") and "Monthly Tabs" in store.get("user_text", ""))

    # chosenLayout without a sheets array is ignored (no build instruction injected)
    store = {}
    imp._generate = gen_returns(json.dumps(CANNED), store)
    _drive(imp, json.dumps({"prompt": "a budget", "chosenLayout": {"title": "Bad"}}).encode())
    check("chosenLayout: malformed (no sheets) is ignored",
          "CHOSEN LAYOUT" not in store.get("user_text", ""))

    # locale folded in (for local-currency recommendations)
    store = {}
    imp._generate = gen_returns(json.dumps(CANNED), store)
    _drive(imp, json.dumps({"prompt": "a price list", "locale": "de-CH"}).encode())
    check("locale: folded into the user message", "User locale: de-CH" in store.get("user_text", ""))

    # attachments reach the provider (validated) + a note is added to the prompt
    store = {}
    imp._generate = gen_returns(json.dumps(CANNED), store)
    _drive(imp, json.dumps({"prompt": "expenses from this receipt",
                            "files": [{"type": "image", "media_type": "image/jpeg", "data": img_b64, "name": "r.jpg"}]}).encode())
    check("files: passed to provider",
          isinstance(store.get("files"), list) and len(store["files"]) == 1 and store["files"][0]["kind"] == "image")
    check("files: note added to prompt", "attached file" in store.get("user_text", ""))

    # _validate_files behaviour
    clean = imp._validate_files([{"type": "image", "media_type": "image/jpeg", "data": img_b64, "name": "x.jpg"}])
    check("validate: image kept", len(clean) == 1 and clean[0]["kind"] == "image" and clean[0]["media_type"] == "image/jpeg")
    check("validate: pdf kept",
          imp._validate_files([{"type": "pdf", "media_type": "application/pdf", "data": img_b64}])[0]["kind"] == "pdf")
    check("validate: none -> []", imp._validate_files(None) == [])

    def rejects_files(label, files):
        try:
            imp._validate_files(files)
            check(label + " (should reject)", False)
        except imp._InputError:
            check(label, True)

    rejects_files("validate: reject too many", [{"type": "image", "media_type": "image/png", "data": "x"}] * (imp.MAX_FILES + 1))
    rejects_files("validate: reject bad image type", [{"type": "image", "media_type": "image/tiff", "data": "x"}])
    rejects_files("validate: reject pdf with image type", [{"type": "pdf", "media_type": "image/png", "data": "x"}])
    rejects_files("validate: reject unknown kind", [{"type": "video", "media_type": "video/mp4", "data": "x"}])
    rejects_files("validate: reject empty data", [{"type": "image", "media_type": "image/png", "data": ""}])
    rejects_files("validate: reject oversized", [{"type": "image", "media_type": "image/png", "data": "a" * (imp.MAX_FILE_B64 + 1)}])

    # rate limit -> graceful 503
    imp._generate = gen_raises("rate_limit")
    status, parsed, _ = _drive(imp, json.dumps({"prompt": "x"}).encode())
    check("rate_limit: graceful 503", status == 503 and "usage limit" in (parsed or {}).get("error", "").lower())

    # auth / quota -> graceful 503
    imp._generate = gen_raises("auth")
    status, parsed, _ = _drive(imp, json.dumps({"prompt": "x"}).encode())
    check("auth: graceful 503", status == 503 and "temporarily unavailable" in (parsed or {}).get("error", "").lower())

    # malformed model output on BOTH the first call and the corrective retry -> 502
    imp._generate = gen_returns("not json at all")
    status, parsed, _ = _drive(imp, json.dumps({"prompt": "x"}).encode())
    check("malformed (both tries): 502", status == 502 and "malformed" in (parsed or {}).get("error", "").lower())

    # malformed first, valid on the corrective retry -> 200 (the retry recovers it)
    seq = {"n": 0}
    def gen_retry(system, user_text, files):
        seq["n"] += 1
        return "garbage, not json" if seq["n"] == 1 else json.dumps(CANNED)
    imp._generate = gen_retry
    status, parsed, _ = _drive(imp, json.dumps({"prompt": "x"}).encode())
    check("retry: malformed then valid -> 200", status == 200 and (parsed or {}).get("status") == "ready")
    check("retry: the corrective retry actually fired", seq["n"] == 2)

    # a spec with a stray non-array row ("null" string) is repaired, not passed through
    BAD_ROW = {"status": "ready", "improvedPrompt": "p", "notes": "n",
               "spec": {"title": "B", "sheets": [{"name": "S",
                        "columns": [{"header": "Cat", "type": "text"}, {"header": "Amt", "type": "number"}],
                        "rows": [["Rent", 100], "null", ["Food", 50]]}]}}
    imp._generate = gen_returns(json.dumps(BAD_ROW))
    status, parsed, _ = _drive(imp, json.dumps({"prompt": "budget"}).encode())
    rows = (((parsed or {}).get("spec") or {}).get("sheets") or [{}])[0].get("rows", [])
    check("sanitize: 200 with the bad row dropped",
          status == 200 and len(rows) == 2 and all(isinstance(r, list) for r in rows))

    # a "layouts" response is returned to the client unchanged
    LAYOUTS = {"status": "layouts", "notes": "pick one",
               "layouts": [{"title": "Flat", "summary": "one sheet", "sheets": [{"name": "S", "columns": ["A", "B"]}]},
                           {"title": "Tabs", "summary": "many", "sheets": [{"name": "Jan", "columns": ["A"]}]}]}
    imp._generate = gen_returns(json.dumps(LAYOUTS))
    status, parsed, _ = _drive(imp, json.dumps({"prompt": "a budget"}).encode())
    check("layouts: 200 passthrough",
          status == 200 and (parsed or {}).get("status") == "layouts" and len((parsed or {}).get("layouts", [])) == 2)

    # directBuild (Regenerate) instructs the model to skip the layout step
    store = {}
    imp._generate = gen_returns(json.dumps(CANNED), store)
    _drive(imp, json.dumps({"prompt": "a refined budget", "directBuild": True}).encode())
    check("directBuild: instructs a direct build (skip layouts)",
          "do NOT propose layouts" in store.get("user_text", ""))

    # directBuild is suppressed while editing (an edit always builds anyway)
    store = {}
    imp._generate = gen_returns(json.dumps(CANNED), store)
    _drive(imp, json.dumps({"prompt": "x", "directBuild": True,
                            "baseSpec": {"title": "T", "sheets": [{"name": "S",
                                         "columns": [{"header": "A", "type": "text"}]}]}}).encode())
    check("directBuild: suppressed during an edit", "do NOT propose layouts" not in store.get("user_text", ""))

    # _sanitize_spec unit behaviour
    san = imp._sanitize_spec
    check("sanitize_spec: drops non-array rows",
          san({"sheets": [{"columns": [{"header": "A"}], "rows": [["x"], "null", 3]}]})["sheets"][0]["rows"] == [["x"]])
    check("sanitize_spec: no sheets -> None", san({"sheets": []}) is None)
    check("sanitize_spec: sheet without columns dropped -> None", san({"sheets": [{"name": "S"}]}) is None)
    check("sanitize_spec: not a dict -> None", san("nope") is None)

    # _normalize_result unit behaviour
    norm = imp._normalize_result
    check("normalize: ready without a spec -> None", norm({"status": "ready", "notes": "n"}) is None)
    check("normalize: needs_input without questions -> None", norm({"status": "needs_input"}) is None)
    check("normalize: layouts with options passes",
          (norm({"status": "layouts", "layouts": [{"title": "A", "sheets": []}]}) or {}).get("status") == "layouts")
    check("normalize: layouts empty -> None", norm({"status": "layouts", "layouts": []}) is None)
    check("normalize: missing status but valid spec -> ready",
          (norm({"spec": {"sheets": [{"columns": [{"header": "A"}], "rows": []}]}}) or {}).get("status") == "ready")

    # _wants_live_data: grounding (scarce quota) only fires for current/external-data prompts
    wld = imp._wants_live_data
    check("live-data: everyday budget -> no grounding", wld("A monthly budget to track my income and expenses") is False)
    check("live-data: meal plan -> no grounding", wld("A weekly meal plan with a shopping list") is False)
    check("live-data: invoice -> no grounding", wld("A simple invoice for my freelance work") is False)
    check("live-data: latest iPhone prices -> grounding", wld("Compare the latest iPhone models with their prices") is True)
    check("live-data: crypto prices -> grounding", wld("Current bitcoin and ethereum prices") is True)

    # _extract_json robustness (prompt-JSON envelope parsing)
    ej = imp._extract_json
    check("extract: plain envelope", (ej('{"status":"ready","notes":"n"}') or {}).get("status") == "ready")
    check("extract: code-fenced", (ej('```json\n{"status":"ready","notes":"n"}\n```') or {}).get("status") == "ready")
    check("extract: stray leading brace + prose",
          (ej('use the {row} token, then: {"status":"ready","notes":"n"}') or {}).get("status") == "ready")
    check("extract: prefers the status envelope among two objects",
          (ej('{"foo":1} then {"status":"needs_input","notes":"n"}') or {}).get("status") == "needs_input")
    check("extract: pure garbage -> None", ej("no json here at all") is None)

    # canQueue surfaces only when a backup worker (PI_WORKER_URL) is configured
    imp._generate = gen_raises("rate_limit")
    _pw = imp.PI_WORKER_URL
    imp.PI_WORKER_URL = None
    _, p_off, _ = _drive(imp, json.dumps({"prompt": "x"}).encode())
    check("canQueue: absent when no worker configured", not (p_off or {}).get("canQueue"))
    imp.PI_WORKER_URL = "https://pi.example.com"
    _, p_on, _ = _drive(imp, json.dumps({"prompt": "x"}).encode())
    check("canQueue: present when worker configured", (p_on or {}).get("canQueue") is True)
    imp.PI_WORKER_URL = _pw


def test_queue():
    print("\n[queue.py]")
    import urllib.request as _ur
    import urllib.error as _ue
    q = _load("queue_mod", os.path.join("api", "queue.py"))

    # not configured -> 503
    q.PI_WORKER_URL = None
    status, _, _ = _drive(q, json.dumps({"prompt": "x", "email": "a@b.co"}).encode())
    check("queue: no worker configured -> 503", status == 503)

    q.PI_WORKER_URL = "https://pi.example.com"
    status, _, _ = _drive(q, json.dumps({"email": "a@b.co"}).encode())
    check("queue: missing prompt -> 400", status == 400)
    status, _, _ = _drive(q, json.dumps({"prompt": "x", "email": "not-an-email"}).encode())
    check("queue: invalid email -> 400", status == 400)

    class _R:
        def __init__(self, st): self.status = st
        def __enter__(self): return self
        def __exit__(self, *a): return False

    captured = {}
    def _ok(req, timeout=None):
        captured["url"] = req.full_url
        captured["secret"] = req.headers.get("X-worker-secret")
        return _R(202)

    _orig = _ur.urlopen
    _ur.urlopen = _ok
    q.PI_WORKER_SECRET = "s3cret"
    try:
        status, parsed, _ = _drive(q, json.dumps({"prompt": "budget", "email": "a@b.co", "data": "x"}).encode())
        check("queue: success -> 200 queued", status == 200 and (parsed or {}).get("queued") is True)
        check("queue: forwards to /generate-async", str(captured.get("url", "")).endswith("/generate-async"))
        check("queue: sends the shared secret header", captured.get("secret") == "s3cret")

        def _boom(req, timeout=None): raise _ue.URLError("down")
        _ur.urlopen = _boom
        status, _, _ = _drive(q, json.dumps({"prompt": "x", "email": "a@b.co"}).encode())
        check("queue: worker unreachable -> 504", status == 504)
    finally:
        _ur.urlopen = _orig


if __name__ == "__main__":
    test_generate()
    test_improve()
    test_queue()
    print("\n==== %d passed, %d failed ====" % (PASSED, FAILED))
    sys.exit(1 if FAILED else 0)
