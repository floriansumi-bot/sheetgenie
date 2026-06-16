"""SheetGenie stress-test harness — runs prompts through the REAL pipeline
(/api/improve live -> /api/generate render) and grades the produced .xlsx.

Derived from spreadsheet_generator_tests_combined.pdf. Each case grades
structurally + by data fidelity (feature presence, correct cell types, formula
shape, preserved values) — robust without needing a formula recalculator.

Reads the API key from .env (never printed). Costs ~3-8c per case on Opus.

Usage:
  .venv\\Scripts\\python.exe tests\\stress_suite.py            # default representative subset
  .venv\\Scripts\\python.exe tests\\stress_suite.py all         # every defined case
  .venv\\Scripts\\python.exe tests\\stress_suite.py A_totals B_computed   # named cases
"""
import importlib.util
import io
import json
import os
import sys
import datetime
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _load(name, rel):
    spec = importlib.util.spec_from_file_location(name, os.path.join(ROOT, rel))
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


def _load_dotenv():
    p = os.path.join(ROOT, ".env")
    if not os.path.isfile(p):
        return
    for line in io.open(p, encoding="utf-8"):
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            k, v = k.strip(), v.strip().strip('"').strip("'")
            if k and k not in os.environ:
                os.environ[k] = v


imp = _load("improve_mod", os.path.join("api", "improve.py"))
gen = _load("generate_mod", os.path.join("api", "generate.py"))


# --------------------------------------------------------------------------
# Pipeline helpers
# --------------------------------------------------------------------------
def improve(body):
    raw = json.dumps(body).encode("utf-8")
    cap = {"status": None}

    class T(imp.handler):
        def __init__(self):
            self.rfile = io.BytesIO(raw)
            self.wfile = io.BytesIO()
            self.headers = {"Content-Length": str(len(raw))}

        def send_response(self, c):
            cap["status"] = c

        def send_header(self, k, v):
            pass

        def end_headers(self):
            pass

    h = T()
    h.do_POST()
    try:
        return cap["status"], json.loads(h.wfile.getvalue().decode("utf-8"))
    except Exception:
        return cap["status"], None


def build(spec):
    gen._validate_spec(spec)            # raises SpecError -> caught by caller
    return load_workbook(io.BytesIO(gen._build_workbook(spec)))


# --------------------------------------------------------------------------
# Grading helpers (openpyxl)
# --------------------------------------------------------------------------
def headers(ws):
    return [c.value for c in ws[1]]


def data_rows(ws):
    out = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v not in (None, "") for v in vals):
            out.append(vals)
    return out


def all_formulas(wb):
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.data_type == "f" and isinstance(c.value, str):
                    out.append(c.value)
    return out


def any_formula(wb, *substrs):
    fs = all_formulas(wb)
    return any(all(s.lower() in f.lower() for s in substrs) for f in fs)


def number_formats(wb):
    s = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.number_format:
                    s.add(c.number_format)
    return s


def has_condfmt(ws):
    try:
        return len(list(ws.conditional_formatting)) > 0
    except Exception:
        return False


def has_validation(ws):
    try:
        return len(ws.data_validations.dataValidation) > 0
    except Exception:
        return False


def has_named(wb):
    try:
        return len(list(wb.defined_names)) > 0
    except Exception:
        return False


PASS, PARTIAL, FAIL = "PASS", "PARTIAL", "FAIL"


# --------------------------------------------------------------------------
# Test cases  (id, prompt, optional data, expect, grade)
# --------------------------------------------------------------------------
def g_totals(wb, spec, resp):
    if any_formula(wb, "=sum("):
        return PASS, "live =SUM totals row present"
    return FAIL, "no =SUM totals formula found"


def g_multisheet(wb, spec, resp):
    names = [ws.title.lower() for ws in wb.worksheets]
    got = sum(1 for q in ("q1", "q2", "q3", "q4") if any(q in n for n in names))
    return (PASS if got == 4 else (PARTIAL if got >= 2 else FAIL)), "quarter sheets found: %d (%s)" % (got, names)


def g_crosssheet(wb, spec, resp):
    return (PASS if any_formula(wb, "!") else FAIL), "cross-sheet ref " + ("present" if any_formula(wb, "!") else "missing")


def g_currency(wb, spec, resp):
    fmts = number_formats(wb)
    ok = any(("chf" in f.lower()) or ("[$" in f) or ("€" in f) or ("$" in f and "#" in f) for f in fmts)
    return (PASS if ok else PARTIAL), "currency-ish formats: " + str([f for f in fmts if any(x in f.lower() for x in ("chf", "$", "€"))][:4])


def g_percent(wb, spec, resp):
    fmts = number_formats(wb)
    pct = any("%" in f for f in fmts)
    div = any_formula(wb, "/")
    if pct and div:
        return PASS, "percent format + division formula"
    if pct:
        return PARTIAL, "percent format but no division formula"
    return FAIL, "no percent formatting"


def g_condfmt(wb, spec, resp):
    ok = any(has_condfmt(ws) for ws in wb.worksheets)
    return (PASS if ok else FAIL), "conditional-formatting rule " + ("present" if ok else "missing")


def g_validation(wb, spec, resp):
    ok = any(has_validation(ws) for ws in wb.worksheets)
    return (PASS if ok else FAIL), "list validation " + ("present" if ok else "missing")


def g_chart(wb, spec, resp):
    ok = any(len(ws._charts) > 0 for ws in wb.worksheets)
    return (PASS if ok else FAIL), "chart " + ("present" if ok else "missing")


def g_if(wb, spec, resp):
    return (PASS if any_formula(wb, "if(") else FAIL), "IF formula " + ("present" if any_formula(wb, "if(") else "missing")


def g_named(wb, spec, resp):
    return (PASS if has_named(wb) else FAIL), "named range " + ("present" if has_named(wb) else "missing")


def g_unicode(wb, spec, resp):
    want = ["Café", "Résumé", "Größe", "日本語"]
    hs = []
    for ws in wb.worksheets:
        hs += [str(h) for h in headers(ws) if h]
    got = [w for w in want if any(w in h for h in hs)]
    return (PASS if len(got) == 4 else PARTIAL), "unicode headers preserved: %d/4" % len(got)


def g_verbatim(wb, spec, resp):
    ws = wb.worksheets[0]
    rows = data_rows(ws)
    h = [str(x).lower() for x in headers(ws) if x]
    amount_col = next((i for i, x in enumerate(h) if "amount" in x), None)
    nums_ok = amount_col is not None and all(isinstance(r[amount_col], (int, float)) for r in rows[:5] if len(r) > amount_col)
    return (PASS if (len(rows) == 5 and nums_ok) else (PARTIAL if len(rows) == 5 else FAIL)), \
        "rows=%d, amounts numeric=%s" % (len(rows), nums_ok)


def g_computed(wb, spec, resp):
    prod = any_formula(wb, "*")            # Total = Qty * UnitPrice
    total = any_formula(wb, "=sum(")       # grand-total row
    if prod and total:
        return PASS, "per-row product formula + grand-total SUM"
    if prod:
        return PARTIAL, "product formula but no grand-total SUM"
    return FAIL, "no product formula"


def g_dedup(wb, spec, resp):
    ws = wb.worksheets[0]
    rows = data_rows(ws)
    return (PASS if len(rows) == 4 else (PARTIAL if len(rows) in (4, 5) else FAIL)), "unique data rows=%d (want 4)" % len(rows)


def g_aggregation(wb, spec, resp):
    # Either literal regional totals (2950/2080/760) appear, or SUMIF formulas do.
    flat = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                flat.append(c.value)
    literals = {2950, 2080, 760}
    got = {v for v in flat if isinstance(v, (int, float)) and v in literals}
    if got == literals:
        return PASS, "literal regional totals present (2950/2080/760)"
    if any_formula(wb, "sumif"):
        return PARTIAL, "SUMIF present (values not literal; recalc to confirm)"
    return FAIL, "regional totals not found"


def g_filter(wb, spec, resp):
    ws = wb.worksheets[0]
    rows = data_rows(ws)
    return (PASS if len(rows) == 3 else (PARTIAL if len(rows) <= 4 else FAIL)), "rows kept=%d (want 3)" % len(rows)


CASES = [
    {"id": "A_totals", "tag": "A", "prompt": "Create a table of 4 products with monthly sales for Jan, Feb, Mar and Apr (numbers), and add a totals row at the bottom of every numeric column.", "grade": g_totals},
    {"id": "A_multisheet", "tag": "A", "prompt": "Create one sheet per quarter named Q1, Q2, Q3 and Q4, each with Revenue and Expenses columns and 3 example rows.", "grade": g_multisheet},
    {"id": "A_crosssheet", "tag": "A", "prompt": "Create sheets Q1, Q2, Q3, Q4 each with a Revenue column and 3 rows, then a Summary sheet that pulls each quarter's total revenue by referencing the quarter sheets.", "grade": g_crosssheet},
    {"id": "A_currency", "tag": "A", "prompt": "Make a simple invoice with 4 line items, amounts priced in Swiss francs (CHF).", "grade": g_currency},
    {"id": "A_percent", "tag": "A", "prompt": "List 5 expense categories with amounts and add a column showing each expense as a percent of the total.", "grade": g_percent},
    {"id": "A_condfmt", "tag": "A", "prompt": "List 8 expenses with amounts and highlight any expense over 500 in red.", "grade": g_condfmt},
    {"id": "A_validation", "tag": "A", "prompt": "Create a task list with a Status column that has a dropdown allowing only: To Do, In Progress, Done.", "grade": g_validation},
    {"id": "A_chart", "tag": "A", "prompt": "Create monthly revenue for 6 months and add a bar chart of the revenue.", "grade": g_chart},
    {"id": "A_if", "tag": "A", "prompt": "List 6 students with a Score column and add a Result column that shows Pass if the score is 60 or higher, otherwise Fail.", "grade": g_if},
    {"id": "A_named", "tag": "A", "prompt": "Create an invoice that uses a named range for the VAT tax rate of 8.1% and references that named range in the VAT formula.", "grade": g_named},
    {"id": "A_unicode", "tag": "A", "prompt": "Create a table with columns named exactly Café, Résumé, Größe and 日本語, with 2 example rows.", "grade": g_unicode},
    {"id": "A_ambiguous", "tag": "A", "prompt": "Make me a tracker.", "expect": "needs_input", "grade": None},

    {"id": "B_verbatim", "tag": "B",
     "data": "OrderID,Customer,Item,Amount\n1001,Mara Vogt,Notebook,12.50\n1002,Tomas Keller,Pen set,8.90\n1003,Aicha Diallo,Desk lamp,34.00\n1004,Lena Brun,Stapler,6.75\n1005,Pavel Roth,Monitor,189.00",
     "prompt": "Put this into a spreadsheet exactly, one column per field.", "grade": g_verbatim},
    {"id": "B_computed", "tag": "B",
     "data": "Product,Qty,UnitPrice\nMug,4,9.50\nT-shirt,3,19.90\nSticker pack,10,2.50\nPoster,2,14.00",
     "prompt": "Add a Total column = Qty x UnitPrice (as a formula), then a grand-total row.", "grade": g_computed},
    {"id": "B_aggregation", "tag": "B",
     "data": "Region,Rep,Sales\nWest,Anya,1200\nEast,Bram,980\nWest,Carlos,1450\nNorth,Dina,760\nEast,Bram,1100\nWest,Anya,300",
     "prompt": "Summarise total sales by region.", "grade": g_aggregation},
    {"id": "B_dedup", "tag": "B",
     "data": "ID,Name,Email\n7,Nora Frei,nora@ex.com\n8,Leo Marti,leo@ex.com\n7,Nora Frei,nora@ex.com\n9,Sara Ott,sara@ex.com\n8,Leo Marti,leo@ex.com\n10,Tim Bopp,tim@ex.com",
     "prompt": "Remove exact duplicate rows.", "grade": g_dedup},
    {"id": "B_filter", "tag": "B",
     "data": "Order,Amount_CHF\nA-1,420\nA-2,1500\nA-3,990\nA-4,2300\nA-5,1001\nA-6,75",
     "prompt": "Keep only orders over 1,000 CHF.", "grade": g_filter},
]

DEFAULT = ["A_totals", "A_multisheet", "A_currency", "A_percent", "A_condfmt", "A_validation",
           "A_chart", "A_if", "A_named", "A_ambiguous", "B_verbatim", "B_computed", "B_dedup", "B_aggregation"]


def run_case(c):
    body = {"prompt": c["prompt"], "hasData": bool(c.get("data")), "data": c.get("data")}
    status, resp = improve(body)
    if resp is None:
        return FAIL, "no/!json response (HTTP %s)" % status
    if c.get("expect") == "needs_input":
        return (PASS if resp.get("status") == "needs_input" else FAIL), "status=%s" % resp.get("status")
    if resp.get("status") == "needs_input":
        clar = [{"question": q.get("question"), "answer": "Use sensible defaults"} for q in resp.get("questions", [])]
        body["clarifications"] = clar
        status, resp = improve(body)
        if resp is None:
            return FAIL, "no response after clarifications"
    spec = resp.get("spec")
    if not spec:
        return FAIL, "no spec (status %s)" % resp.get("status")
    try:
        wb = build(spec)
    except gen.SpecError as e:
        return FAIL, "generate rejected the spec: %s" % e
    except Exception as e:  # noqa: BLE001
        return FAIL, "render error: %s" % e
    try:
        return c["grade"](wb, spec, resp)
    except Exception as e:  # noqa: BLE001
        return FAIL, "grader error: %s" % e


def main():
    _load_dotenv()
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ANTHROPIC_API_KEY not set (put it in .env).")
        return 2
    args = sys.argv[1:]
    if args == ["all"]:
        ids = [c["id"] for c in CASES]
    elif args:
        ids = args
    else:
        ids = DEFAULT
    by_id = {c["id"]: c for c in CASES}

    results = []
    print("Running %d stress cases (model: %s)\n" % (len(ids), imp.MODEL_CHAIN[0]))
    for tid in ids:
        c = by_id.get(tid)
        if not c:
            print("  ?     %s (unknown case)" % tid)
            continue
        try:
            verdict, detail = run_case(c)
        except Exception as e:  # noqa: BLE001
            verdict, detail = FAIL, "exception: %s" % e
        results.append({"id": tid, "verdict": verdict, "detail": detail})
        mark = {"PASS": "PASS", "PARTIAL": "PART", "FAIL": "FAIL"}[verdict]
        print("  %-4s  %-14s  %s" % (mark, tid, detail))

    n = len(results)
    p = sum(1 for r in results if r["verdict"] == PASS)
    pa = sum(1 for r in results if r["verdict"] == PARTIAL)
    f = sum(1 for r in results if r["verdict"] == FAIL)
    print("\n==== %d pass, %d partial, %d fail / %d ====" % (p, pa, f, n))
    out = os.path.join(ROOT, "_stress_report.local.json")
    with io.open(out, "w", encoding="utf-8") as fh:
        json.dump(results, fh, indent=2, ensure_ascii=False)
    print("report:", out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
