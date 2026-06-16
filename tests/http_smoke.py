"""HTTP smoke test against a running dev server (tools/devserver.py on :8123).

Exercises the full pipeline over real HTTP: GET /, POST /api/improve (mock spec),
POST /api/generate (REAL openpyxl render) -> validates the returned .xlsx bytes.
Run the dev server first, then:  .venv\\Scripts\\python.exe tests\\http_smoke.py [port]
"""
import io
import json
import sys
import urllib.request
from openpyxl import load_workbook

PORT = sys.argv[1] if len(sys.argv) > 1 else "8123"
BASE = "http://127.0.0.1:%s" % PORT


def post(path, obj):
    req = urllib.request.Request(
        BASE + path, data=json.dumps(obj).encode("utf-8"),
        headers={"Content-Type": "application/json"}, method="POST")
    return urllib.request.urlopen(req, timeout=20)


def main():
    # GET /
    root = urllib.request.urlopen(BASE + "/", timeout=10).read().decode("utf-8", "ignore")
    assert "SheetGenie" in root, "index did not contain SheetGenie"
    print("GET / -> 200, has SheetGenie")

    # POST /api/improve
    data = json.loads(post("/api/improve", {"prompt": "budget tracker"}).read())
    spec = data["spec"]
    assert "improvedPrompt" in data and spec.get("sheets"), "improve shape wrong"
    print("POST /api/improve -> 200, title:", spec.get("title"))

    # POST /api/generate (real render) -> xlsx
    resp = post("/api/generate", {"spec": spec, "filename": "smoke"})
    ctype = resp.headers.get("Content-Type", "")
    disp = resp.headers.get("Content-Disposition", "")
    body = resp.read()
    assert "spreadsheetml.sheet" in ctype, "wrong content-type: " + ctype
    assert "attachment" in disp and ".xlsx" in disp, "bad disposition: " + disp
    assert len(body) > 2000, "xlsx too small"

    wb = load_workbook(io.BytesIO(body))
    ws = wb["Budget"]
    assert ws["D2"].data_type == "f", "D2 not a formula"
    assert len(ws._charts) == 1, "expected one chart"
    print("POST /api/generate -> 200, %d bytes, D2=%s, charts=%d, disp=%s"
          % (len(body), ws["D2"].value, len(ws._charts), disp))

    print("HTTP SMOKE PASS")


if __name__ == "__main__":
    main()
