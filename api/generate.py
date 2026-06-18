"""
SheetGenie — POST /api/generate

Pure, deterministic rendering of a SpreadsheetSpec (see docs/SPEC.md) into a real
.xlsx workbook using openpyxl. NO AI, NO network calls. Vercel Python serverless
function using the native http.server.BaseHTTPRequestHandler pattern.

Contract (docs/SPEC.md §1):
  Request body : { "spec": <SpreadsheetSpec>, "filename": "string|null" }
  Success 200  : raw .xlsx bytes, attachment download headers
  Error 4xx/5xx: JSON { "error": "<safe human-readable message>" }
"""

import json
import re
import sys
import datetime
import traceback
from io import BytesIO
from http.server import BaseHTTPRequestHandler

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, Rule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


# ---------------------------------------------------------------------------
# Limits (docs/SPEC.md §2 "Hard limits")
# ---------------------------------------------------------------------------
MAX_SHEETS = 8
MIN_SHEETS = 1
MAX_COLUMNS = 50
MIN_COLUMNS = 1
MAX_ROWS = 5000
MIN_ROWS = 0
MAX_CHARTS = 6
MIN_CHARTS = 0
MAX_VALUE_COLUMNS = 10
MIN_VALUE_COLUMNS = 1
MAX_CONDITIONAL_FORMATS = 20
MIN_CONDITIONAL_FORMATS = 0
MAX_DATA_VALIDATIONS = 20
MIN_DATA_VALIDATIONS = 0
MAX_VALIDATION_VALUES = 200
MIN_VALIDATION_VALUES = 1
MAX_NAMED_RANGES = 50
MIN_NAMED_RANGES = 0

# Abuse / resource caps (defense-in-depth; the body cap is the primary guard).
MAX_BODY_BYTES = 4 * 1024 * 1024   # 4 MB request body
MAX_CELL_CHARS = 32767             # Excel's hard per-cell character limit
MAX_FORMULA_CHARS = 8192           # Excel's formula length limit
MAX_HEADER_CHARS = 255
MAX_TOTAL_CELLS = 200_000          # across the whole workbook

VALID_COLUMN_TYPES = {"text", "number", "currency", "percent", "date", "formula"}
VALID_CHART_TYPES = {"bar", "line", "pie"}

# Column types that get a live =SUM in a totals row.
SUMMABLE_TYPES = {"number", "currency", "percent", "formula"}

# Conditional-format rule names (docs/SPEC.md §2).
COMPARISON_RULES = {
    "greaterThan", "greaterThanOrEqual", "lessThan", "lessThanOrEqual",
    "equal", "between",
}
VALID_CONDFMT_RULES = COMPARISON_RULES | {"top10", "bottom10", "colorScale"}

# Friendly colour names -> standard light fill hex (no leading alpha).
CONDFMT_COLORS = {
    "red": "FFC7CE",
    "green": "C6EFCE",
    "yellow": "FFEB9C",
    "orange": "FFD8A8",
    "blue": "BDD7EE",
}
_HEX6 = re.compile(r"^[0-9A-Fa-f]{6}$")

# 3-colour scale endpoints for colorScale rules (light red -> yellow -> green).
COLORSCALE_MIN = "F8696B"
COLORSCALE_MID = "FFEB84"
COLORSCALE_MAX = "63BE7B"

# Column type -> Excel number format (docs/SPEC.md §2 table).
TYPE_FORMATS = {
    "text": "General",
    "number": "#,##0.00",
    "currency": "#,##0.00",
    "percent": "0.0%",
    "date": "yyyy-mm-dd",
    # "formula" inherits from neighbours -> no forced format.
}

HEADER_FILL = PatternFill(start_color="FFE8EEF7", end_color="FFE8EEF7", fill_type="solid")
HEADER_FONT = Font(bold=True)

TOTAL_FONT = Font(bold=True)
TOTAL_TOP_BORDER = Border(top=Side(style="thin"))

MIN_COL_WIDTH = 8
MAX_COL_WIDTH = 48
WIDTH_PADDING = 2        # breathing room beyond the widest cell in a column
FILTER_ARROW_PAD = 3     # extra header room so the autofilter dropdown doesn't cover it


class SpecError(ValueError):
    """Raised for any spec validation failure -> 400."""


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate_spec(spec):
    """Validate a SpreadsheetSpec against the hard limits in SPEC.md.

    Raises SpecError (-> 400) with a clear reason on any failure.
    """
    if not isinstance(spec, dict):
        raise SpecError("spec must be an object.")

    sheets = spec.get("sheets")
    if not isinstance(sheets, list):
        raise SpecError("spec.sheets must be a list.")
    if not (MIN_SHEETS <= len(sheets) <= MAX_SHEETS):
        raise SpecError(
            "spec must have between %d and %d sheets (got %d)."
            % (MIN_SHEETS, MAX_SHEETS, len(sheets))
        )

    total_cells = 0

    for si, sheet in enumerate(sheets):
        where = "sheet %d" % (si + 1)
        if not isinstance(sheet, dict):
            raise SpecError("%s must be an object." % where)

        columns = sheet.get("columns")
        if not isinstance(columns, list):
            raise SpecError("%s: columns must be a list." % where)
        if not (MIN_COLUMNS <= len(columns) <= MAX_COLUMNS):
            raise SpecError(
                "%s: must have between %d and %d columns (got %d)."
                % (where, MIN_COLUMNS, MAX_COLUMNS, len(columns))
            )
        n_cols = len(columns)

        for ci, col in enumerate(columns):
            cwhere = "%s, column %d" % (where, ci + 1)
            if not isinstance(col, dict):
                raise SpecError("%s must be an object." % cwhere)
            ctype = col.get("type")
            if ctype not in VALID_COLUMN_TYPES:
                raise SpecError(
                    "%s: invalid type %r (allowed: %s)."
                    % (cwhere, ctype, ", ".join(sorted(VALID_COLUMN_TYPES)))
                )
            header = col.get("header")
            if isinstance(header, str) and len(header) > MAX_HEADER_CHARS:
                raise SpecError("%s: header is too long." % cwhere)
            if ctype == "formula":
                formula = col.get("formula")
                if not isinstance(formula, str) or not formula.strip():
                    raise SpecError(
                        "%s: type 'formula' requires a non-empty 'formula' string."
                        % cwhere
                    )
                if len(formula) > MAX_FORMULA_CHARS:
                    raise SpecError("%s: formula is too long." % cwhere)

        rows = sheet.get("rows")
        if rows is None:
            rows = []
        if not isinstance(rows, list):
            raise SpecError("%s: rows must be a list." % where)
        if not (MIN_ROWS <= len(rows) <= MAX_ROWS):
            raise SpecError(
                "%s: must have between %d and %d rows (got %d)."
                % (where, MIN_ROWS, MAX_ROWS, len(rows))
            )

        # Per-cell string-length cap (Excel's own limit) — bounds individual
        # cell size even within a shape-valid spec.
        for ri, row in enumerate(rows):
            if not isinstance(row, list):
                raise SpecError("%s: row %d must be a list." % (where, ri + 1))
            for cell in row:
                if isinstance(cell, str) and len(cell) > MAX_CELL_CHARS:
                    raise SpecError(
                        "%s: a cell value exceeds the %d-character limit."
                        % (where, MAX_CELL_CHARS)
                    )

        # Total-cell budget across the whole workbook (resource guard).
        total_cells += n_cols * len(rows)
        if total_cells > MAX_TOTAL_CELLS:
            raise SpecError(
                "The spreadsheet is too large (over %d cells). "
                "Please request fewer rows or columns." % MAX_TOTAL_CELLS
            )

        charts = sheet.get("charts")
        if charts is None:
            charts = []
        if not isinstance(charts, list):
            raise SpecError("%s: charts must be a list." % where)
        if not (MIN_CHARTS <= len(charts) <= MAX_CHARTS):
            raise SpecError(
                "%s: must have between %d and %d charts (got %d)."
                % (where, MIN_CHARTS, MAX_CHARTS, len(charts))
            )

        for chi, chart in enumerate(charts):
            chwhere = "%s, chart %d" % (where, chi + 1)
            if not isinstance(chart, dict):
                raise SpecError("%s must be an object." % chwhere)
            if chart.get("type") not in VALID_CHART_TYPES:
                raise SpecError(
                    "%s: invalid chart type %r (allowed: %s)."
                    % (chwhere, chart.get("type"), ", ".join(sorted(VALID_CHART_TYPES)))
                )

            cat_col = chart.get("categoriesColumn")
            if not isinstance(cat_col, int) or isinstance(cat_col, bool):
                raise SpecError("%s: categoriesColumn must be an integer." % chwhere)
            if not (1 <= cat_col <= n_cols):
                raise SpecError(
                    "%s: categoriesColumn %d out of range (1..%d)."
                    % (chwhere, cat_col, n_cols)
                )

            value_cols = chart.get("valueColumns")
            if not isinstance(value_cols, list):
                raise SpecError("%s: valueColumns must be a list." % chwhere)
            if not (MIN_VALUE_COLUMNS <= len(value_cols) <= MAX_VALUE_COLUMNS):
                raise SpecError(
                    "%s: valueColumns must have between %d and %d entries (got %d)."
                    % (chwhere, MIN_VALUE_COLUMNS, MAX_VALUE_COLUMNS, len(value_cols))
                )
            for vc in value_cols:
                if not isinstance(vc, int) or isinstance(vc, bool):
                    raise SpecError("%s: valueColumns entries must be integers." % chwhere)
                if not (1 <= vc <= n_cols):
                    raise SpecError(
                        "%s: valueColumns entry %d out of range (1..%d)."
                        % (chwhere, vc, n_cols)
                    )

            # Optional row bounds, if given, must be sane integers.
            for key in ("dataStartRow", "dataEndRow"):
                val = chart.get(key)
                if val is None:
                    continue
                if not isinstance(val, int) or isinstance(val, bool) or val < 1:
                    raise SpecError("%s: %s must be a positive integer." % (chwhere, key))

        # --- totalsRow (optional bool) ---
        totals = sheet.get("totalsRow")
        if totals is not None and not isinstance(totals, bool):
            raise SpecError("%s: totalsRow must be a boolean." % where)

        # --- conditionalFormats (optional list) ---
        cond_fmts = sheet.get("conditionalFormats")
        if cond_fmts is None:
            cond_fmts = []
        if not isinstance(cond_fmts, list):
            raise SpecError("%s: conditionalFormats must be a list." % where)
        if not (MIN_CONDITIONAL_FORMATS <= len(cond_fmts) <= MAX_CONDITIONAL_FORMATS):
            raise SpecError(
                "%s: must have between %d and %d conditionalFormats (got %d)."
                % (where, MIN_CONDITIONAL_FORMATS, MAX_CONDITIONAL_FORMATS, len(cond_fmts))
            )
        for cfi, cf in enumerate(cond_fmts):
            cfwhere = "%s, conditionalFormat %d" % (where, cfi + 1)
            if not isinstance(cf, dict):
                raise SpecError("%s must be an object." % cfwhere)
            col = cf.get("column")
            if not isinstance(col, int) or isinstance(col, bool):
                raise SpecError("%s: column must be an integer." % cfwhere)
            if not (1 <= col <= n_cols):
                raise SpecError(
                    "%s: column %d out of range (1..%d)." % (cfwhere, col, n_cols)
                )
            rule = cf.get("rule")
            if rule not in VALID_CONDFMT_RULES:
                raise SpecError(
                    "%s: invalid rule %r (allowed: %s)."
                    % (cfwhere, rule, ", ".join(sorted(VALID_CONDFMT_RULES)))
                )
            if rule in COMPARISON_RULES:
                value = cf.get("value")
                if not isinstance(value, (int, float, str)) or isinstance(value, bool):
                    raise SpecError(
                        "%s: rule %r requires a numeric or string 'value'."
                        % (cfwhere, rule)
                    )
                if rule == "between":
                    value2 = cf.get("value2")
                    if not isinstance(value2, (int, float, str)) or isinstance(value2, bool):
                        raise SpecError(
                            "%s: rule 'between' requires a numeric or string 'value2'."
                            % cfwhere
                        )
            # Colour (used by comparison + top10/bottom10; colorScale ignores it).
            color = cf.get("color")
            if color is not None:
                if not isinstance(color, str) or _resolve_fill_hex(color) is None:
                    raise SpecError(
                        "%s: invalid color %r (use red/green/yellow/orange/blue "
                        "or a 6-hex)." % (cfwhere, color)
                    )

        # --- dataValidations (optional list) ---
        validations = sheet.get("dataValidations")
        if validations is None:
            validations = []
        if not isinstance(validations, list):
            raise SpecError("%s: dataValidations must be a list." % where)
        if not (MIN_DATA_VALIDATIONS <= len(validations) <= MAX_DATA_VALIDATIONS):
            raise SpecError(
                "%s: must have between %d and %d dataValidations (got %d)."
                % (where, MIN_DATA_VALIDATIONS, MAX_DATA_VALIDATIONS, len(validations))
            )
        for dvi, dv in enumerate(validations):
            dvwhere = "%s, dataValidation %d" % (where, dvi + 1)
            if not isinstance(dv, dict):
                raise SpecError("%s must be an object." % dvwhere)
            col = dv.get("column")
            if not isinstance(col, int) or isinstance(col, bool):
                raise SpecError("%s: column must be an integer." % dvwhere)
            if not (1 <= col <= n_cols):
                raise SpecError(
                    "%s: column %d out of range (1..%d)." % (dvwhere, col, n_cols)
                )
            values = dv.get("values")
            if not isinstance(values, list):
                raise SpecError("%s: values must be a list." % dvwhere)
            if not (MIN_VALIDATION_VALUES <= len(values) <= MAX_VALIDATION_VALUES):
                raise SpecError(
                    "%s: values must have between %d and %d entries (got %d)."
                    % (dvwhere, MIN_VALIDATION_VALUES, MAX_VALIDATION_VALUES, len(values))
                )
            for v in values:
                if not isinstance(v, (int, float, str)) or isinstance(v, bool):
                    raise SpecError(
                        "%s: values entries must be strings or numbers." % dvwhere
                    )
            # Excel's inline list-validation source (formula1) is limited to 255
            # chars; reject loudly rather than emit a workbook Excel will "repair".
            joined = ",".join(str(v).replace(",", " ").replace('"', '""') for v in values)
            if len(joined) + 2 > 255:
                raise SpecError(
                    "%s: the dropdown list is too long; keep the combined options "
                    "under ~250 characters." % dvwhere
                )

    # --- namedRanges (workbook-level, optional list) ---
    named = spec.get("namedRanges")
    if named is None:
        named = []
    if not isinstance(named, list):
        raise SpecError("namedRanges must be a list.")
    if not (MIN_NAMED_RANGES <= len(named) <= MAX_NAMED_RANGES):
        raise SpecError(
            "spec must have between %d and %d namedRanges (got %d)."
            % (MIN_NAMED_RANGES, MAX_NAMED_RANGES, len(named))
        )
    for nri, nr in enumerate(named):
        nwhere = "namedRange %d" % (nri + 1)
        if not isinstance(nr, dict):
            raise SpecError("%s must be an object." % nwhere)
        name = nr.get("name")
        if not isinstance(name, str) or not name.strip():
            raise SpecError("%s: name must be a non-empty string." % nwhere)
        nm = name.strip()
        if (len(nm) > 255
                or not re.match(r"^[A-Za-z_\\][A-Za-z0-9_.]*$", nm)
                or re.match(r"^[A-Za-z]{1,3}[0-9]+$", nm)):
            raise SpecError(
                "%s: invalid defined name %r (start with a letter or underscore, "
                "no spaces, and not a cell reference like A1)." % (nwhere, name)
            )
        ref = nr.get("ref")
        if not isinstance(ref, str) or not ref.strip():
            raise SpecError("%s: ref must be a non-empty string." % nwhere)


# ---------------------------------------------------------------------------
# Rendering helpers
# ---------------------------------------------------------------------------
def _unique_title(name, used):
    """Truncate a sheet name to 31 chars and de-duplicate against `used`."""
    if not isinstance(name, str) or not name.strip():
        base = "Sheet"
    else:
        base = name.strip()
    base = base[:31]
    title = base
    n = 2
    while title in used or not title:
        # Reserve room for the suffix so the result stays <= 31 chars.
        suffix = " (%d)" % n
        title = (base[: 31 - len(suffix)] + suffix) if base else ("Sheet%d" % n)
        n += 1
    used.add(title)
    return title


def _measure_cell(value, ctype):
    """Estimate the displayed character width of one rendered cell value, so columns
    can be sized to their actual content (not just the header)."""
    if value is None:
        return 0
    if isinstance(value, bool):
        return 5  # TRUE / FALSE
    if ctype in ("number", "currency"):
        if isinstance(value, (int, float)):
            s = "{:,.2f}".format(float(value))
            return len(s) + (4 if ctype == "currency" else 0)  # room for a symbol
        return len(str(value))
    if ctype == "percent":
        if isinstance(value, (int, float)):
            return len("{:,.1f}%".format(float(value) * 100))
        return len(str(value))
    if ctype == "date":
        if isinstance(value, str):
            t = value.strip()
            return 10 if len(t) >= 8 else len(t)  # ISO yyyy-mm-dd renders ~10
        return 10
    return len(str(value))


def _formula_cell_width(col):
    """Estimate a calculated (formula) column's value width from its number format —
    the actual value is computed by Excel and unknown at generation time."""
    fmt = col.get("format")
    if isinstance(fmt, str) and "%" in fmt:
        return 8
    if isinstance(fmt, str) and ("0.00" in fmt or "#,##0" in fmt):
        return 14
    return 12


def _auto_widths(columns, rows, autofilter):
    """Content-aware column sizing. Returns (widths, wrap_cols): one width per column
    measured from the header AND the real data (clamped to [MIN, MAX]), plus the set
    of text-column indices whose content is too long to fit and should WRAP instead of
    stretching the column off-screen. Stops measuring a column once it hits MAX."""
    n = len(columns)
    arrow = FILTER_ARROW_PAD if autofilter else 0
    widths = []
    for col in columns:
        header = col.get("header")
        hlen = (len(header) if isinstance(header, str) else 0) + arrow
        if col.get("type") == "formula":
            hlen = max(hlen, _formula_cell_width(col))
        widths.append(hlen)

    capped = [w >= MAX_COL_WIDTH for w in widths]
    overflow = [False] * n  # data content exceeded the cap -> wrap candidate
    for row in rows:
        if all(capped):
            break
        if not isinstance(row, list):
            continue
        for ci in range(n):
            if capped[ci] or columns[ci].get("type") == "formula":
                continue
            val = row[ci] if ci < len(row) else None
            w = _measure_cell(val, columns[ci].get("type"))
            if w > widths[ci]:
                widths[ci] = w
                if w >= MAX_COL_WIDTH:
                    capped[ci] = True
                    overflow[ci] = True

    final = [max(MIN_COL_WIDTH, min(MAX_COL_WIDTH, w + WIDTH_PADDING)) for w in widths]
    wrap_cols = {ci for ci in range(n) if overflow[ci] and columns[ci].get("type") == "text"}
    return final, wrap_cols


def _coerce_date(value):
    """Coerce an ISO string to a date/datetime object; otherwise return as-is."""
    if isinstance(value, str):
        s = value.strip()
        try:
            return datetime.date.fromisoformat(s)
        except ValueError:
            pass
        try:
            return datetime.datetime.fromisoformat(s)
        except ValueError:
            pass
    return value


def _resolve_fill_hex(color):
    """Map a friendly colour name or 6-hex string to an ARGB hex, else None.

    Accepts the named colours (red/green/yellow/orange/blue) and any raw
    6-digit hex (e.g. "FFC7CE"). Returns an 8-digit ARGB string suitable for
    PatternFill, or None if the input is not a recognised colour.
    """
    if not isinstance(color, str):
        return None
    key = color.strip().lower()
    if key in CONDFMT_COLORS:
        return "FF" + CONDFMT_COLORS[key]
    raw = color.strip()
    if _HEX6.match(raw):
        return "FF" + raw.upper()
    return None


def _render_sheet(ws, sheet):
    """Render one Sheet spec onto an openpyxl worksheet."""
    columns = sheet["columns"]
    rows = sheet.get("rows") or []
    n_cols = len(columns)
    auto_filter_on = sheet.get("autoFilter", True)

    # Content-aware column sizing: measure the header AND the actual data so columns
    # aren't left too wide (blank gaps) or too narrow (clipped text). Long-text
    # columns are capped and wrapped rather than stretched off-screen.
    auto_widths, wrap_cols = _auto_widths(columns, rows, auto_filter_on)

    # --- Headers (row 1) ---
    for ci, col in enumerate(columns, start=1):
        header = col.get("header")
        header = header if isinstance(header, str) else ""
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Column width: honour an explicit spec width ONLY if it is sane — the model
        # sometimes emits absurd values (e.g. 150-250), which produce huge blank
        # columns / sideways scroll. Anything out of [MIN, MAX] falls back to the
        # measured content width.
        width = col.get("width")
        letter = get_column_letter(ci)
        if (isinstance(width, (int, float)) and not isinstance(width, bool)
                and MIN_COL_WIDTH <= width <= MAX_COL_WIDTH):
            ws.column_dimensions[letter].width = float(width)
        else:
            ws.column_dimensions[letter].width = auto_widths[ci - 1]

    # --- Data rows (start at row 2) ---
    first_data_row = 2
    last_data_row = first_data_row + len(rows) - 1  # == 1 when rows is empty

    for ri, row in enumerate(rows):
        actual_row = first_data_row + ri
        # Pad/truncate the row to exactly len(columns).
        cells = list(row) if isinstance(row, list) else []
        if len(cells) < n_cols:
            cells = cells + [None] * (n_cols - len(cells))
        else:
            cells = cells[:n_cols]

        for ci, col in enumerate(columns, start=1):
            ctype = col.get("type")
            target = ws.cell(row=actual_row, column=ci)

            if ctype == "formula":
                # Formula columns ignore any value supplied in rows.
                formula = col.get("formula") or ""
                target.value = formula.replace("{row}", str(actual_row))
            else:
                value = cells[ci - 1]
                if ctype == "date":
                    value = _coerce_date(value)
                target.value = value
                # Formula-injection defense: openpyxl turns a leading "=" string
                # into a live formula. In a non-formula column, force any such
                # cell back to a literal string so user data can never execute.
                if target.data_type == "f":
                    target.data_type = "s"

            # Number format: explicit column.format overrides the type default.
            explicit = col.get("format")
            if isinstance(explicit, str) and explicit:
                target.number_format = explicit
            else:
                fmt = TYPE_FORMATS.get(ctype)
                if fmt:
                    target.number_format = fmt

            if ctype == "text":
                if (ci - 1) in wrap_cols:
                    target.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    target.alignment = Alignment(horizontal="left")

    # --- Freeze header (default true) ---
    if sheet.get("freezeHeader", True):
        ws.freeze_panes = "A2"

    # --- Auto filter (default true) ---
    if auto_filter_on:
        end_col = get_column_letter(n_cols)
        end_row = last_data_row if rows else 1
        ws.auto_filter.ref = "A1:%s%d" % (end_col, end_row)

    # --- Charts (skip entirely if there are no data rows) ---
    # Charts must reference the DATA range only, so render them before the
    # totals row is appended below the data.
    if rows:
        _render_charts(ws, sheet, columns, first_data_row, last_data_row)

    # --- Conditional formats + data validations on the DATA range ---
    # (Both apply to firstDataRow..lastDataRow; they skip cleanly with no rows.)
    if rows:
        _render_conditional_formats(ws, sheet, columns, first_data_row, last_data_row)
    _render_data_validations(ws, sheet, columns, first_data_row, last_data_row, bool(rows))

    # --- Totals row (appended below the data; excluded from filter + charts) ---
    if rows and sheet.get("totalsRow"):
        _render_totals_row(ws, columns, first_data_row, last_data_row)


def _render_charts(ws, sheet, columns, first_data_row, last_data_row):
    n_cols = len(columns)
    charts = sheet.get("charts") or []
    for chart_spec in charts:
        ctype = chart_spec["type"]
        if ctype == "bar":
            chart = BarChart()
        elif ctype == "line":
            chart = LineChart()
        else:  # "pie"
            chart = PieChart()

        chart.title = chart_spec.get("title") or None

        cat_col = chart_spec["categoriesColumn"]
        data_start = chart_spec.get("dataStartRow") or first_data_row
        data_end = chart_spec.get("dataEndRow") or last_data_row
        # Clamp to the populated range so an over-specified bound can't extend
        # the series over empty cells (trailing zero/blank points).
        data_start = max(data_start, first_data_row)
        data_end = min(data_end, last_data_row)
        if data_end < data_start:
            continue  # nothing to plot

        categories = Reference(
            ws, min_col=cat_col, min_row=data_start, max_row=data_end
        )

        value_cols = chart_spec["valueColumns"]
        if ctype == "pie":
            value_cols = value_cols[:1]  # pie uses only the first value column

        for vc in value_cols:
            # The value series row span MUST equal the categories span
            # (data_start..data_end) or values and labels misalign whenever
            # dataStartRow > the first data row. Take the series name from the
            # header cell explicitly rather than folding row 1 into the range.
            series_ref = Reference(ws, min_col=vc, min_row=data_start, max_row=data_end)
            chart.add_data(series_ref, titles_from_data=False)
            header_val = ws.cell(row=1, column=vc).value
            if header_val is not None and chart.series:
                try:
                    chart.series[-1].tx = SeriesLabel(v=str(header_val))
                except Exception:  # noqa: BLE001 — title is cosmetic; never fail the chart
                    pass

        chart.set_categories(categories)

        anchor = chart_spec.get("anchor")
        if not (isinstance(anchor, str) and anchor):
            anchor = "%s2" % get_column_letter(n_cols + 2)
        ws.add_chart(chart, anchor)


def _render_totals_row(ws, columns, first_data_row, last_data_row):
    """Append one totals row below the data.

    Each number/currency/percent/formula column gets a live
    =SUM(<col><first>:<col><last>) over the DATA range with that column's
    number format. The first text column (or column 1 if there is none) gets a
    bold "Total" label. The whole row is bold with a thin top border.
    """
    total_row = last_data_row + 1
    n_cols = len(columns)

    # Pick the label column: first text column, else column 1.
    label_col = 1
    for ci, col in enumerate(columns, start=1):
        if col.get("type") == "text":
            label_col = ci
            break

    for ci, col in enumerate(columns, start=1):
        ctype = col.get("type")
        target = ws.cell(row=total_row, column=ci)
        letter = get_column_letter(ci)

        if ctype in SUMMABLE_TYPES:
            target.value = "=SUM(%s%d:%s%d)" % (
                letter, first_data_row, letter, last_data_row
            )
            # Match the column's number format (explicit overrides type).
            explicit = col.get("format")
            if isinstance(explicit, str) and explicit:
                target.number_format = explicit
            else:
                fmt = TYPE_FORMATS.get(ctype)
                if fmt:
                    target.number_format = fmt
        elif ci == label_col:
            target.value = "Total"

        # Bold + thin top border across the whole row.
        target.font = TOTAL_FONT
        target.border = TOTAL_TOP_BORDER


def _render_conditional_formats(ws, sheet, columns, first_data_row, last_data_row):
    """Apply real openpyxl conditional formatting to columns' data ranges."""
    cond_fmts = sheet.get("conditionalFormats") or []
    for cf in cond_fmts:
        col = cf["column"]
        letter = get_column_letter(col)
        cell_range = "%s%d:%s%d" % (letter, first_data_row, letter, last_data_row)
        rule_name = cf["rule"]

        if rule_name in COMPARISON_RULES:
            fill = PatternFill(
                start_color=_resolve_fill_hex(cf.get("color")) or ("FF" + CONDFMT_COLORS["yellow"]),
                end_color=_resolve_fill_hex(cf.get("color")) or ("FF" + CONDFMT_COLORS["yellow"]),
                fill_type="solid",
            )
            value = cf.get("value")
            if rule_name == "between":
                lo, hi = value, cf.get("value2")
                # Excel's "between" needs lo <= hi; swap reversed numeric bounds so
                # the rule highlights rather than silently matching nothing.
                if isinstance(lo, (int, float)) and isinstance(hi, (int, float)) and lo > hi:
                    lo, hi = hi, lo
                formula = [_condfmt_operand(lo), _condfmt_operand(hi)]
            else:
                formula = [_condfmt_operand(value)]
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator=rule_name, formula=formula, fill=fill),
            )

        elif rule_name in ("top10", "bottom10"):
            fill = PatternFill(
                start_color=_resolve_fill_hex(cf.get("color")) or ("FF" + CONDFMT_COLORS["green"]),
                end_color=_resolve_fill_hex(cf.get("color")) or ("FF" + CONDFMT_COLORS["green"]),
                fill_type="solid",
            )
            dxf = DifferentialStyle(fill=fill)
            rule = Rule(
                type="top10",
                rank=10,
                percent=True,
                bottom=(rule_name == "bottom10"),
                dxf=dxf,
            )
            ws.conditional_formatting.add(cell_range, rule)

        else:  # "colorScale"
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="min", start_color=COLORSCALE_MIN,
                    mid_type="percentile", mid_value=50, mid_color=COLORSCALE_MID,
                    end_type="max", end_color=COLORSCALE_MAX,
                ),
            )


def _condfmt_operand(value):
    """Format a comparison operand for a CellIsRule formula.

    Numbers pass through as their literal text; strings are wrapped in double
    quotes so Excel compares against the literal text.
    """
    if isinstance(value, bool):
        return '"%s"' % value
    if isinstance(value, (int, float)):
        return repr(value) if isinstance(value, float) else str(value)
    return '"%s"' % str(value).replace('"', '""')


def _render_data_validations(ws, sheet, columns, first_data_row, last_data_row, has_rows):
    """Attach real list (dropdown) data validations to columns' data ranges."""
    validations = sheet.get("dataValidations") or []
    if not validations:
        return
    # With no data rows there is no range to attach to; skip cleanly.
    if not has_rows:
        return

    for dv_spec in validations:
        col = dv_spec["column"]
        letter = get_column_letter(col)
        cell_range = "%s%d:%s%d" % (letter, first_data_row, letter, last_data_row)

        # formula1 is a quoted, comma-joined list. Commas are the list separator,
        # so replace them with a space; double embedded quotes so a value containing
        # a " cannot break out of the literal into a live formula. (The combined
        # length is capped to Excel's 255-char list-source limit in _validate_spec.)
        items = [str(v).replace(",", " ").replace('"', '""') for v in dv_spec["values"]]
        joined = ",".join(items)
        formula1 = '"%s"' % joined

        dv = DataValidation(
            type="list",
            formula1=formula1,
            allow_blank=True,
            showDropDown=False,        # False => the dropdown arrow IS shown
            showErrorMessage=True,
            showInputMessage=True,
            errorStyle="stop",
            errorTitle="Invalid entry",
            error="Pick a value from the list.",
            promptTitle="Choose a value",
            prompt="Select one of the allowed values.",
        )
        ws.add_data_validation(dv)
        dv.add(cell_range)


def _render_named_ranges(wb, spec):
    """Register workbook-level defined names (openpyxl 3.1.5 API)."""
    named = spec.get("namedRanges") or []
    for nr in named:
        name = nr["name"].strip()
        ref = nr["ref"].strip()
        wb.defined_names.add(DefinedName(name=name, attr_text=ref))


def _slugify_filename(filename, spec):
    """Build a safe download filename: slug.xlsx."""
    candidate = None
    if isinstance(filename, str) and filename.strip():
        candidate = filename
    elif isinstance(spec, dict) and isinstance(spec.get("title"), str) and spec["title"].strip():
        candidate = spec["title"]
    else:
        candidate = "spreadsheet"

    slug = candidate.strip().lower()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    slug = re.sub(r"-+", "-", slug).strip("-")
    if not slug:
        slug = "spreadsheet"
    return slug + ".xlsx"


def _build_workbook(spec):
    """Render a validated SpreadsheetSpec into .xlsx bytes."""
    wb = Workbook()
    used_titles = set()

    for idx, sheet in enumerate(spec["sheets"]):
        if idx == 0:
            ws = wb.active  # reuse the default sheet for the first
        else:
            ws = wb.create_sheet()
        ws.title = _unique_title(sheet.get("name"), used_titles)
        _render_sheet(ws, sheet)

    # Workbook-level defined names (must be added after sheets exist).
    _render_named_ranges(wb, spec)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# HTTP handler
# ---------------------------------------------------------------------------
class handler(BaseHTTPRequestHandler):
    def _send_json(self, status, payload):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_error(self, status, message):
        self._send_json(status, {"error": message})

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Content-Length", "0")
        self.end_headers()

    def do_POST(self):
        try:
            # --- Read + parse body ---
            try:
                length = int(self.headers.get("Content-Length") or 0)
            except (TypeError, ValueError):
                length = 0
            if length > MAX_BODY_BYTES:
                self._send_error(413, "Request body is too large.")
                return
            raw = self.rfile.read(min(length, MAX_BODY_BYTES)) if length > 0 else b""

            try:
                payload = json.loads(raw.decode("utf-8")) if raw else None
            except (ValueError, UnicodeDecodeError):
                self._send_error(400, "Request body is not valid JSON.")
                return

            if not isinstance(payload, dict):
                self._send_error(400, "Request body must be a JSON object.")
                return

            spec = payload.get("spec")
            filename = payload.get("filename")

            # --- Validate spec ---
            try:
                _validate_spec(spec)
            except SpecError as e:
                self._send_error(400, str(e))
                return

            # --- Render ---
            try:
                data = _build_workbook(spec)
            except SpecError as e:
                self._send_error(400, str(e))
                return

            safe_name = _slugify_filename(filename, spec)

            # --- Respond with the .xlsx bytes ---
            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header(
                "Content-Disposition", 'attachment; filename="%s"' % safe_name
            )
            self.send_header("Content-Length", str(len(data)))
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(data)

        except Exception:  # noqa: BLE001 — last-resort safety net
            traceback.print_exc(file=sys.stderr)
            try:
                self._send_error(
                    500, "Failed to generate the spreadsheet. Please try again."
                )
            except Exception:  # noqa: BLE001 — headers may already be sent
                traceback.print_exc(file=sys.stderr)
